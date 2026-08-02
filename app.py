from flask import Flask, render_template, request, redirect, session, flash, send_file
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter
from reportlab.platypus import (
    SimpleDocTemplate,
    Table,
    TableStyle,
    Paragraph,
    Spacer
)
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.pdfgen import canvas
import sqlite3
import os
from openpyxl import Workbook
from datetime import datetime

app = Flask(__name__)
app.secret_key = "finsight_secret_key"


# ---------------- DATABASE ----------------

conn = sqlite3.connect("database.db")
cursor = conn.cursor()

cursor.execute("""
CREATE TABLE IF NOT EXISTS savings(
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    user_id INTEGER,
    savings_name TEXT,
    amount REAL,
    date TEXT
)
""")

conn.commit()
conn.close()


# ---------------- LOGIN ----------------

@app.route("/", methods=["GET", "POST"])
def login():

    if request.method == "POST":

        email = request.form["email"]
        password = request.form["password"]

        conn = sqlite3.connect("database.db")
        cursor = conn.cursor()

        cursor.execute(
            "SELECT * FROM users WHERE email=? AND password=?",
            (email, password)
        )

        user = cursor.fetchone()

        conn.close()

        if user:
            session["user_id"] = user[0]
            return redirect("/dashboard")
        else:
            flash("Invalid Email or Password")

    return render_template("login.html")
@app.route("/register", methods=["GET", "POST"])
def register():

    if request.method == "POST":

        name = request.form["name"]
        email = request.form["email"]
        password = request.form["password"]

        conn = sqlite3.connect("database.db")
        cursor = conn.cursor()

        cursor.execute(
            "SELECT * FROM users WHERE email=?",
            (email,)
        )

        existing = cursor.fetchone()

        if existing:
            flash("Email already registered")
            conn.close()
            return redirect("/register")

        cursor.execute("""
            INSERT INTO users(name,email,password)
            VALUES(?,?,?)
        """, (name, email, password))

        conn.commit()
        conn.close()

        flash("Registration Successful")
        return redirect("/")

    return render_template("register.html")

# ---------------- FORGOT PASSWORD ----------------

@app.route("/forgot_password", methods=["GET","POST"])
def forgot_password():

    if request.method=="POST":

        email=request.form["email"]
        new_password=request.form["new_password"]

        conn=sqlite3.connect("database.db")
        cursor=conn.cursor()

        cursor.execute(
            "UPDATE users SET password=? WHERE email=?",
            (new_password,email)
        )

        conn.commit()
        conn.close()

        flash("Password Updated Successfully")
        return redirect("/")

    return render_template("forgot_password.html")
@app.route("/dashboard")
def dashboard():

    user_id = session.get("user_id")

    if not user_id:
        return redirect("/")

    conn = sqlite3.connect("database.db")
    cursor = conn.cursor()

    # ---------------- Total Income ----------------
    cursor.execute(
        "SELECT SUM(amount) FROM income WHERE user_id=?",
        (user_id,)
    )
    total_income = cursor.fetchone()[0] or 0

    # ---------------- Total Expense ----------------
    cursor.execute(
        "SELECT SUM(amount) FROM expenses WHERE user_id=?",
        (user_id,)
    )
    total_expense = cursor.fetchone()[0] or 0

    # ---------------- Expense Category ----------------
    cursor.execute("""
        SELECT expense_name,
               SUM(amount)
        FROM expenses
        WHERE user_id=?
        GROUP BY expense_name
    """, (user_id,))
    expense_chart = cursor.fetchall()

    # ---------------- Monthly Expense Trend ----------------
    cursor.execute("""
        SELECT strftime('%m', date),
               SUM(amount)
        FROM expenses
        WHERE user_id=?
        GROUP BY strftime('%m', date)
        ORDER BY strftime('%m', date)
    """, (user_id,))
    expense_trend = cursor.fetchall()

    # ---------------- Budget ----------------
    cursor.execute(
        "SELECT SUM(amount) FROM budgets WHERE user_id=?",
        (user_id,)
    )
    total_budget = cursor.fetchone()[0] or 0

    # ---------------- Investment ----------------
    cursor.execute("""
        SELECT SUM(invested_amount),
               SUM(current_value)
        FROM investments
        WHERE user_id=?
    """, (user_id,))

    investment = cursor.fetchone()

    total_investment = investment[0] or 0
    current_value = investment[1] or 0

    profit = current_value - total_investment

    if total_investment > 0:
        returns = round((profit / total_investment) * 100, 2)
    else:
        returns = 0
    

    
    # ---------------- Transactions ----------------

    cursor.execute("""
        SELECT id,
               income_name,
               amount,
               date,
               'Income'
        FROM income
        WHERE user_id=?
    """, (user_id,))
    income_data = cursor.fetchall()

    cursor.execute("""
    SELECT id,
           expense_name,
           amount,
           date,
           'Expense',
           payment_mode
    FROM expenses
    WHERE user_id=?
""", (user_id,))
    expense_data = cursor.fetchall()

    transactions = income_data + expense_data
    search = request.args.get("search", "")

    if search:
       transactions = [
        t for t in transactions
        if search.lower() in str(t[1]).lower()
    ]
    transactions.sort(key=lambda x: x[3], reverse=True)

    # ---------------- Financial Health ----------------

    monthly_income = total_income
    monthly_expense = total_expense
    monthly_savings = monthly_income - monthly_expense

    if monthly_income > 0:
        health_score = round((monthly_savings / monthly_income) * 100)
    else:
        health_score = 0

    if health_score < 0:
        health_score = 0

    if health_score >= 80:
        health_status = "Excellent"
    elif health_score >= 60:
        health_status = "Good"
    elif health_score >= 40:
        health_status = "Average"
    else:
        health_status = "Poor"

    # ---------------- Budget Recommendation ----------------

    if total_expense > total_income:
        recommendation = "⚠️ Your expenses are higher than your income. Reduce unnecessary spending."

    elif monthly_savings <= 0:
        recommendation = "⚠️ Try to save some amount every month."

    elif monthly_savings < total_income * 0.20:
        recommendation = "💡 Increase your monthly savings to at least 20% of your income."

    else:
        recommendation = "✅ Excellent! Your financial management is on the right track."

    # ---------------- Notification ----------------

    notifications = []
   
   # Savings Notification
    if monthly_savings > 0:
          notifications.append(f"🎉 Great! You saved ₹{monthly_savings:,.0f} this month.")
    else:
          notifications.append("⚠️ No savings this month.")
   
   # Budget Notification
    if total_budget > 0:
        if total_expense > total_budget:
           exceeded = total_expense - total_budget
           notifications.append(f"🔴 Budget exceeded by ₹{exceeded:,.0f}")
    else:
           remaining = total_budget - total_expense
           notifications.append(f"🟢 Budget remaining ₹{remaining:,.0f}")
   
   # Investment Notification
    if profit > 0:
           notifications.append(f"📈 Investment Profit ₹{profit:,.0f}")
    elif profit < 0:
           notifications.append(f"📉 Investment Loss ₹{abs(profit):,.0f}")
    else:
           notifications.append("➖ No Profit / Loss")
   
   # Health Notification
    notifications.append(f"❤️ Financial Health : {health_status}")
       
   

    conn.close()

    return render_template(
        "dashboard.html",
        total_income=total_income,
        total_expense=total_expense,
        total_budget=total_budget,
        total_investment=total_investment,
        current_value=current_value,
        profit=profit,
        returns=returns,
        transactions=transactions,
        search=search,
        expense_chart=expense_chart,
        expense_trend=expense_trend,
        monthly_income=monthly_income,
        monthly_expense=monthly_expense,
        monthly_savings=monthly_savings,
        health_score=health_score,
        health_status=health_status,
        recommendation=recommendation,
        notifications=notifications,
    )

    
@app.route("/income", methods=["GET", "POST"])
def income():

    user_id = session.get("user_id")

    if not user_id:
        return redirect("/")

    conn = sqlite3.connect("database.db")
    cursor = conn.cursor()

    if request.method == "POST":

        income_name = request.form["income_name"]
        amount = request.form["amount"]
        date = request.form["date"]

        cursor.execute("""
            INSERT INTO income
            (user_id, income_name, amount, date)
            VALUES (?,?,?,?)
        """, (
            user_id,
            income_name,
            amount,
            date
        ))

        conn.commit()
        flash("Income Added Successfully!")

    cursor.execute("""
        SELECT id,
               income_name,
               amount,
               date
        FROM income
        WHERE user_id=?
        ORDER BY date DESC
    """, (user_id,))

    incomes = cursor.fetchall()

    conn.close()

    return render_template(
        "income.html",
        incomes=incomes
    )
@app.route("/edit_income/<int:id>", methods=["GET", "POST"])
def edit_income(id):

    conn = sqlite3.connect("database.db")
    cursor = conn.cursor()

    if request.method == "POST":

        income_name = request.form["income_name"]
        amount = request.form["amount"]
        date = request.form["date"]

        cursor.execute("""
            UPDATE income
            SET income_name=?,
                amount=?,
                date=?
            WHERE id=?
        """, (
            income_name,
            amount,
            date,
            id
        ))

        conn.commit()
        conn.close()

        return redirect("/dashboard")

    cursor.execute("""
        SELECT id,
               income_name,
               amount,
               date
        FROM income
        WHERE id=?
    """, (id,))

    income = cursor.fetchone()

    conn.close()

    return render_template(
        "edit_income.html",
        income=income
    )
@app.route("/delete_income/<int:id>")
def delete_income(id):

    conn = sqlite3.connect("database.db")
    cursor = conn.cursor()

    cursor.execute("DELETE FROM income WHERE id=?", (id,))
    conn.commit()
    conn.close()

    return redirect("/dashboard")
@app.route("/expense", methods=["GET", "POST"])
def expense():

    user_id = session.get("user_id")

    if not user_id:
        return redirect("/")

    conn = sqlite3.connect("database.db")
    cursor = conn.cursor()

    if request.method == "POST":

        expense_name = request.form["expense_name"]
        amount = request.form["amount"]
        date = request.form["date"]
        category = request.form["category"]
        payment_mode = request.form["payment_mode"]
        description = request.form["description"]

        cursor.execute("""
            INSERT INTO expenses
            (
                user_id,
                expense_name,
                amount,
                date,
                category,
                payment_mode,
                description
            )
            VALUES (?,?,?,?,?,?,?)
        """, (
            user_id,
            expense_name,
            amount,
            date,
            category,
            payment_mode,
            description
        ))

        conn.commit()
        flash("Expense Added Successfully!")

    cursor.execute("""
        SELECT id,
               expense_name,
               amount,
               date,
               category,
               payment_mode,
               description
        FROM expenses
        WHERE user_id=?
        ORDER BY date DESC
    """, (user_id,))

    expenses = cursor.fetchall()

    conn.close()

    return render_template(
        "expense.html",
        expenses=expenses
    )
@app.route("/edit_expense/<int:id>", methods=["GET", "POST"])
def edit_expense(id):

    conn = sqlite3.connect("database.db")
    cursor = conn.cursor()

    if request.method == "POST":

        expense_name = request.form["expense_name"]
        amount = request.form["amount"]
        date = request.form["date"]

        cursor.execute("""
        UPDATE expenses
        SET expense_name=?, amount=?, date=?
        WHERE id=?
        """, (expense_name, amount, date, id))

        conn.commit()
        conn.close()

        return redirect("/dashboard")

    cursor.execute("""
    SELECT id, expense_name, amount, date
    FROM expenses
    WHERE id=?
    """, (id,))

    expense = cursor.fetchone()

    conn.close()

    return render_template("edit_expense.html", expense=expense)
@app.route("/delete_expense/<int:id>")
def delete_expense(id):

    conn = sqlite3.connect("database.db")
    cursor = conn.cursor()

    cursor.execute("DELETE FROM expenses WHERE id=?", (id,))
    conn.commit()
    conn.close()

    return redirect("/dashboard")
@app.route("/budget", methods=["GET", "POST"])
def budget():

    user_id = session.get("user_id")

    if not user_id:
        return redirect("/")

    conn = sqlite3.connect("database.db")
    cursor = conn.cursor()

    if request.method == "POST":

        budget_name = request.form["budget_name"]
        amount = request.form["amount"]
        category = request.form["category"]
        start_date = request.form["start_date"]

        cursor.execute("""
            INSERT INTO budgets
            (
                user_id,
                budget_name,
                amount,
                category,
                start_date
            )
            VALUES (?,?,?,?,?)
        """, (
            user_id,
            budget_name,
            amount,
            category,
            start_date
        ))

        conn.commit()
        flash("Budget Added Successfully!")

    cursor.execute("""
        SELECT id,
               budget_name,
               amount,
               category,
               start_date
        FROM budgets
        WHERE user_id=?
        ORDER BY start_date DESC
    """, (user_id,))

    budgets = cursor.fetchall()

    conn.close()

    return render_template(
        "budget.html",
        budgets=budgets
    )
@app.route("/savings")
def savings():

    user_id = session.get("user_id")

    if not user_id:
        return redirect("/")

    conn = sqlite3.connect("database.db")
    cursor = conn.cursor()

    cursor.execute("""
        SELECT id,
               savings_name,
               amount,
               date
        FROM savings
        WHERE user_id=?
        ORDER BY date DESC
    """, (user_id,))

    savings = cursor.fetchall()

    conn.close()

    return render_template(
        "savings.html",
        savings=savings
    )
@app.route("/add_savings", methods=["POST"])
def add_savings():

    user_id = session.get("user_id")

    if not user_id:
        return redirect("/")

    savings_name = request.form["name"]
    amount = request.form["amount"]
    date = request.form["date"]

    conn = sqlite3.connect("database.db")
    cursor = conn.cursor()

    cursor.execute("""
        INSERT INTO savings
        (
            user_id,
            savings_name,
            amount,
            date
        )
        VALUES (?,?,?,?)
    """, (
        user_id,
        savings_name,
        amount,
        date
    ))

    conn.commit()
    conn.close()

    flash("Savings Added Successfully!")

    return redirect("/savings")
@app.route('/edit_savings/<int:id>', methods=['GET', 'POST'])
def edit_savings(id):

    conn = sqlite3.connect('database.db')
    cursor = conn.cursor()

    if request.method == 'POST':

        name = request.form['name']
        amount = request.form['amount']
        date = request.form['date']

        # category remove chesam
        cursor.execute('''
            UPDATE savings
            SET savings_name=?, amount=?, date=?
            WHERE id=?
        ''', (name, amount, date, id))

        conn.commit()
        conn.close()

        return redirect('/savings')

    cursor.execute('SELECT * FROM savings WHERE id=?', (id,))
    saving = cursor.fetchone()

    conn.close()

    return render_template('edit_savings.html', saving=saving)

@app.route('/delete_savings/<int:id>')
def delete_savings(id):
    conn = sqlite3.connect("database.db")
    cursor = conn.cursor()

    cursor.execute("DELETE FROM savings WHERE id=?", (id,))
    conn.commit()
    conn.close()

    return redirect('/savings')
@app.route("/investment", methods=["GET", "POST"])
def investment():

    user_id = session.get("user_id")

    if not user_id:
        return redirect("/")

    conn = sqlite3.connect("database.db")
    cursor = conn.cursor()

    if request.method == "POST":

        investment_name = request.form["investment_name"]
        investment_type = request.form["investment_type"]
        invested_amount = float(request.form["invested_amount"])
        current_value = float(request.form["current_value"])
        investment_date = request.form["investment_date"]

        cursor.execute("""
            INSERT INTO investments
            (
                user_id,
                investment_name,
                investment_type,
                invested_amount,
                current_value,
                investment_date
            )
            VALUES (?,?,?,?,?,?)
        """,(
            user_id,
            investment_name,
            investment_type,
            invested_amount,
            current_value,
            investment_date
        ))

        conn.commit()
        flash("Investment Added Successfully!")

    cursor.execute("""
        SELECT id,
               investment_name,
               investment_type,
               invested_amount,
               current_value,
               investment_date
        FROM investments
        WHERE user_id=?
        ORDER BY investment_date DESC
    """,(user_id,))

    investments = cursor.fetchall()

    total_investment = 0
    total_current = 0

    for i in investments:
        total_investment += float(i[3])
        total_current += float(i[4])

    profit_loss = total_current - total_investment

    if total_investment > 0:
        return_percentage = round((profit_loss / total_investment) * 100,2)
    else:
        return_percentage = 0

    conn.close()

    return render_template(
        "investment.html",
        investments=investments,
        total_investment=total_investment,
        total_current=total_current,
        profit_loss=profit_loss,
        return_percentage=return_percentage
    )
@app.route("/edit_investment/<int:id>", methods=["GET","POST"])
def edit_investment(id):

    conn = sqlite3.connect("database.db")
    cursor = conn.cursor()

    if request.method == "POST":

        cursor.execute("""
        UPDATE investments
        SET investment_name=?,
            investment_type=?,
            invested_amount=?,
            current_value=?,
            investment_date=?
        WHERE id=?
        """,(
            request.form["investment_name"],
            request.form["investment_type"],
            request.form["invested_amount"],
            request.form["current_value"],
            request.form["investment_date"],
            id
        ))

        conn.commit()
        conn.close()

        flash("Investment Updated Successfully!")
        return redirect("/investment")

    cursor.execute("SELECT * FROM investments WHERE id=?",(id,))
    investment = cursor.fetchone()

    conn.close()

    return render_template(
        "edit_investment.html",
        investment=investment
    )
@app.route("/delete_investment/<int:id>")
def delete_investment(id):

    conn = sqlite3.connect("database.db")
    cursor = conn.cursor()

    cursor.execute(
        "DELETE FROM investments WHERE id=?",
        (id,)
    )

    conn.commit()
    conn.close()

    flash("Investment Deleted Successfully!")

    return redirect("/investment")
@app.route("/goal", methods=["GET", "POST"])
def goal():

    user_id = session.get("user_id")

    if not user_id:
        return redirect("/")

    conn = sqlite3.connect("database.db")
    cursor = conn.cursor()

    if request.method == "POST":

        goal_name = request.form["goal_name"]
        target_amount = float(request.form["target_amount"])
        saved_amount = float(request.form["saved_amount"])
        target_date = request.form["target_date"]

        cursor.execute("""
            INSERT INTO goals
            (user_id, goal_name, target_amount, saved_amount, target_date)
            VALUES (?,?,?,?,?)
        """,
        (
            user_id,
            goal_name,
            target_amount,
            saved_amount,
            target_date
        ))

        conn.commit()
        flash("Goal Added Successfully")
        conn.close()
        return redirect("/goal")

    cursor.execute("""
        SELECT id,
               goal_name,
               target_amount,
               saved_amount,
               target_date
        FROM goals
        WHERE user_id=?
    """, (user_id,))

    goals = cursor.fetchall()

    total_target = 0
    total_saved = 0

    for g in goals:
        total_target += float(g[2])
        total_saved += float(g[3])

    conn.close()

    return render_template(
        "goal.html",
        goals=goals,
        total_target=total_target,
        total_saved=total_saved
    )


# ================= EDIT GOAL =================

@app.route("/edit_goal/<int:id>", methods=["GET", "POST"])
def edit_goal(id):

    conn = sqlite3.connect("database.db")
    cursor = conn.cursor()

    if request.method == "POST":

        goal_name = request.form["goal_name"]
        target_amount = request.form["target_amount"]
        saved_amount = request.form["saved_amount"]
        target_date = request.form["target_date"]

        cursor.execute("""
            UPDATE goals
            SET goal_name=?,
                target_amount=?,
                saved_amount=?,
                target_date=?
            WHERE id=?
        """,
        (
            goal_name,
            target_amount,
            saved_amount,
            target_date,
            id
        ))

        conn.commit()
        conn.close()

        flash("Goal Updated Successfully")
        return redirect("/goal")

    cursor.execute(
        "SELECT * FROM goals WHERE id=?",
        (id,)
    )

    goal = cursor.fetchone()

    conn.close()

    return render_template(
        "edit_goal.html",
        goal=goal
    )


# ================= DELETE GOAL =================

@app.route("/delete_goal/<int:id>")
def delete_goal(id):

    conn = sqlite3.connect("database.db")
    cursor = conn.cursor()

    cursor.execute(
        "DELETE FROM goals WHERE id=?",
        (id,)
    )

    conn.commit()
    conn.close()

    flash("Goal Deleted Successfully")

    return redirect("/goal")

@app.route("/analytics")
def analytics():

    user_id = session.get("user_id")

    if not user_id:
        return redirect("/")

    conn = sqlite3.connect("database.db")
    cursor = conn.cursor()

    # Total Income
    cursor.execute(
        "SELECT SUM(amount) FROM income WHERE user_id=?",
        (user_id,)
    )
    total_income = cursor.fetchone()[0] or 0

    # Total Expense
    cursor.execute(
        "SELECT SUM(amount) FROM expenses WHERE user_id=?",
        (user_id,)
    )
    total_expense = cursor.fetchone()[0] or 0
    # Total Budget
    cursor.execute("""
        SELECT SUM(amount)
        FROM budgets
        WHERE user_id=?
    """, (user_id,))

    total_budget = cursor.fetchone()[0] or 0
    goal = cursor.execute(
    "SELECT goal_name, target_amount, saved_amount FROM goals WHERE user_id=?",
    (session["user_id"],)
    ).fetchone()

    cursor.execute("""
       SELECT SUM(invested_amount),
        SUM(current_value)
    FROM investments
    WHERE user_id=?
    """, (user_id,))

    investment = cursor.fetchone()

    total_investment = investment[0] or 0
    current_value = investment[1] or 0
    profit = current_value - total_investment


    # Expense Category Chart
    cursor.execute("""
        SELECT expense_name, SUM(amount)
        FROM expenses
        WHERE user_id=?
        GROUP BY expense_name
    """, (user_id,))
    expense_chart = cursor.fetchall()

    # Monthly Expense Trend
    cursor.execute("""
        SELECT strftime('%m', date) AS month,
               SUM(amount)
        FROM expenses
        WHERE user_id=?
        GROUP BY month
        ORDER BY month
    """, (user_id,))
    expense_trend = cursor.fetchall()

    # Monthly Summary
    monthly_income = total_income
    monthly_expense = total_expense
    monthly_savings = monthly_income - monthly_expense
    # Expense Breakdown
    expense_breakdown = []

    for item in expense_chart:
        category = item[0]
        amount = item[1]

    percent = 0
    if monthly_expense > 0:
        percent = round((amount / monthly_expense) * 100)

    expense_breakdown.append({
        "category": category,
        "amount": amount,
        "percent": percent
    })
    # AI Financial Advisor

    if monthly_income == 0:
       ai_advice = "Please add your income to receive financial advice."

    elif monthly_expense > monthly_income:
       ai_advice = "⚠️ Your expenses are higher than your income. Reduce unnecessary spending."

    elif monthly_savings < monthly_income * 0.2:
       ai_advice = "💡 Try to save at least 20% of your monthly income."

    elif monthly_savings >= monthly_income * 0.5:
       ai_advice = "🎉 Excellent! Your savings are outstanding. Consider investing."

    else:
       ai_advice = "✅ Good financial management. Keep tracking your expenses."

    # Financial Health
    if monthly_income > 0:
        health_score = round((monthly_savings / monthly_income) * 100)
    else:
        health_score = 0

    if health_score < 0:
        health_score = 0

    if health_score >= 80:
        health_status = "Excellent"
    elif health_score >= 60:
        health_status = "Good"
    elif health_score >= 40:
        health_status = "Average"
    else:
        health_status = "Poor"
    # ---------------- Financial Health Metrics ----------------

    if monthly_income > 0:
        spending_score = max(0, 100 - round((monthly_expense / monthly_income) * 100))
        savings_score = round((monthly_savings / monthly_income) * 100)
    else:
        spending_score = 0
        savings_score = 0

# Investment Score
    if total_investment > 0:
        investment_score = round((current_value / total_investment) * 100)
    else:
        investment_score = 0

# Budget Score
    if total_budget > 0:
        budget_score = max(0, 100 - round((total_expense / total_budget) * 100))
    else:
        budget_score = 0

# Limit 0–100
    spending_score = min(max(spending_score, 0), 100)
    savings_score = min(max(savings_score, 0), 100)
    investment_score = min(max(investment_score, 0), 100)
    budget_score = min(max(budget_score, 0), 100)

    # Recommendation
    if total_expense > total_income:
        recommendation = "⚠️ Your expenses are higher than your income. Reduce unnecessary spending."
    elif monthly_savings <= 0:
        recommendation = "⚠️ Try to save some amount every month."
    elif monthly_savings < (0.2 * total_income):
        recommendation = "💡 Increase your monthly savings to at least 20% of your income."
    else:
        recommendation = "✅ Excellent! Your financial management is on the right track."
    # Budget Alert

    remaining_budget = total_budget - total_expense

    if total_budget == 0:
        budget_alert = "No budget has been created."

    elif total_expense >= total_budget:
        budget_alert = "🚨 Budget Limit Exceeded!"

    elif total_expense >= total_budget * 0.8:
       budget_alert = "⚠️ You have used more than 80% of your budget."

    else:
       budget_alert = "✅ Your budget is under control."
    if goal:
        goal_name = goal[0]
        target_amount = goal[1]
        saved_amount = goal[2]
    else:
       goal_name = "No Goal"
       target_amount = 0
       saved_amount = 0

    goal_percent = 0
    if target_amount > 0:
        goal_percent = round((saved_amount / target_amount) * 100)
    if total_investment == 0:
        investment_alert = "No investments added."
    if target_amount > 0:
        goal_percent = round((saved_amount / target_amount) * 100)
    else:
        goal_percent = 0

    remaining_goal = target_amount - saved_amount

    if remaining_goal < 0:
       remaining_goal = 0

    elif profit > 0:
        investment_alert = f"📈 Great! Your investments are in profit by ₹{profit:.2f}."

    elif profit < 0:
        investment_alert = f"📉 Your investments are in loss by ₹{abs(profit):.2f}."

    else:
        investment_alert = "➖ No profit or loss in your investments."
    # Monthly Financial Summary

    monthly_summary = f"""
    Monthly Income : ₹{monthly_income}

    Monthly Expense : ₹{monthly_expense}

    Monthly Savings : ₹{monthly_savings}

    Financial Health : {health_status}

    """
    # Previous Month Data (Demo)
    last_income = 180000
    last_expense = 52000
    last_savings = last_income - last_expense

# Comparison %
    income_change = round(((monthly_income - last_income) / last_income) * 100, 1)
    expense_change = round(((monthly_expense - last_expense) / last_expense) * 100, 1)
    savings_change = round(((monthly_savings - last_savings) / last_savings) * 100, 1)
    insights = []

# Savings comparison
    if monthly_income > 0:
        savings_percent = round((monthly_savings / monthly_income) * 100)

    if savings_percent >= 20:
        insights.append(f"📈 Great! You saved {savings_percent}% of your monthly income.")
    else:
        insights.append(f"💡 Try saving at least 20% of your monthly income.")

# Expense check
    if monthly_expense > (monthly_income * 0.8):
        insights.append("⚠️ Your expenses are very high this month.")
    else:
        insights.append("✅ Your expenses are under control.")

# Investment
    if profit > 0:
        insights.append("📈 Investment portfolio is performing well.")
    else:
        insights.append("⚠️ Your investments are currently in loss.")

# Budget
    if remaining_budget > 0:
        insights.append("✅ You are within your budget limit.")
    else:
        insights.append("🚨 You have exceeded your budget.")
    

    
    conn.close()

    return render_template(
        "analytics.html",
        monthly_income=monthly_income,
        monthly_expense=monthly_expense,
        monthly_savings=monthly_savings,
        expense_chart=expense_chart,
        expense_trend=expense_trend,
        health_score=health_score,
        health_status=health_status,
        recommendation=recommendation,
        ai_advice=ai_advice,
        total_budget=total_budget,
        remaining_budget=remaining_budget,
        budget_alert=budget_alert,
        goal_name=goal_name,
        target_amount=target_amount,
        saved_amount=saved_amount,
        remaining_goal=remaining_goal,
        goal_percent=goal_percent,
        total_investment=total_investment,
        current_value=current_value,
        profit=profit,
        investment_alert=investment_alert,
        monthly_summary=monthly_summary,
        expense_breakdown=expense_breakdown,
        income_change=income_change,
        expense_change=expense_change,
        savings_change=savings_change,
        insights=insights,
        spending_score=spending_score,
        savings_score=savings_score,
        investment_score=investment_score,
        budget_score=budget_score,
    )
@app.route("/financial-health")
def financial_health():

    if "user_id" not in session:
        return redirect("/")

    user_id = session["user_id"]

    conn = sqlite3.connect("database.db")
    cursor = conn.cursor()

    # ---------------- Total Income ----------------

    cursor.execute(
        "SELECT SUM(amount) FROM income WHERE user_id=?",
        (user_id,)
    )
    total_income = cursor.fetchone()[0] or 0

    # ---------------- Total Expense ----------------

    cursor.execute(
        "SELECT SUM(amount) FROM expenses WHERE user_id=?",
        (user_id,)
    )
    total_expense = cursor.fetchone()[0] or 0

    # ---------------- Total Budget ----------------

    cursor.execute(
        "SELECT SUM(amount) FROM budgets WHERE user_id=?",
        (user_id,)
    )
    total_budget = cursor.fetchone()[0] or 0

    # ---------------- Investments ----------------

    cursor.execute("""
        SELECT
            SUM(invested_amount),
            SUM(current_value)
        FROM investments
        WHERE user_id=?
    """, (user_id,))

    investment = cursor.fetchone()

    total_investment = investment[0] or 0
    current_value = investment[1] or 0

    profit = current_value - total_investment

    # ---------------- Calculations ----------------

    total_savings = total_income - total_expense

    # Expense %
    if total_income > 0:
        expense_percent = round((total_expense / total_income) * 100)
    else:
        expense_percent = 0

    # Savings %
    if total_income > 0:
        savings_percent = round((total_savings / total_income) * 100)
    else:
        savings_percent = 0

    # Budget Usage
    if total_budget > 0:
        actual_budget_used = round((total_expense / total_budget) * 100)

        # Progress bar max 100%
        budget_used = min(actual_budget_used, 100)

    else:
        actual_budget_used = 0
        budget_used = 0

    # ---------------- Health Score ----------------

    health_score = savings_percent

    if health_score > 100:
        health_score = 100

    if health_score < 0:
        health_score = 0

    if health_score >= 80:
        health_status = "Excellent"

    elif health_score >= 60:
        health_status = "Good"

    elif health_score >= 40:
        health_status = "Average"

    else:
        health_status = "Poor"

    # ---------------- Investment Status ----------------

    if profit > 0:
        investment_status = "Profit"

    elif profit < 0:
        investment_status = "Loss"

    else:
        investment_status = "No Profit / Loss"

    # ---------------- Financial Tips ----------------

    tips = []

    # Savings Tip
    if savings_percent >= 50:
        tips.append("🟢 Excellent savings habit.")
    else:
        tips.append("🟡 Try to increase your monthly savings.")

    # Expense Tip
    if expense_percent <= 50:
        tips.append("🟢 Expenses are under control.")
    else:
        tips.append("🟡 Your expenses are relatively high. Try reducing unnecessary spending.")

    # Budget Tip
    if total_budget > 0:

        if total_expense > total_budget:

            exceeded = total_expense - total_budget

            budget_message = f"Budget exceeded by ₹{exceeded:,.0f}"

            tips.append(f"🔴 {budget_message}")

        else:

            remaining = total_budget - total_expense

            budget_message = f"₹{remaining:,.0f} budget remaining"

            tips.append(f"🟢 {budget_message}")

    else:

        budget_message = "No budget created"

        tips.append("🟡 Create a monthly budget to track expenses.")

    conn.close()

    return render_template(

        "financial_health.html",

        total_income=total_income,
        total_expense=total_expense,
        total_savings=total_savings,

        total_budget=total_budget,

        expense_percent=expense_percent,
        savings_percent=savings_percent,

        budget_used=budget_used,
        actual_budget_used=actual_budget_used,
        budget_message=budget_message,

        health_score=health_score,
        health_status=health_status,

        total_investment=total_investment,
        current_value=current_value,
        profit=profit,
        investment_status=investment_status,

        tips=tips
    )
@app.route("/budget_recommendation")
def budget_recommendation():

    if "user_id" not in session:
        return redirect("/")

    conn = sqlite3.connect("database.db")
    cursor = conn.cursor()

    user_id = session["user_id"]

    cursor.execute("SELECT SUM(amount) FROM income WHERE user_id=?", (user_id,))
    income = cursor.fetchone()[0] or 0

    cursor.execute("SELECT SUM(amount) FROM expenses WHERE user_id=?", (user_id,))
    expense = cursor.fetchone()[0] or 0

    cursor.execute("SELECT SUM(amount) FROM budgets WHERE user_id=?", (user_id,))
    budget = cursor.fetchone()[0] or 0

    savings = income - expense

    conn.close()

    return render_template(
        "budget_recommendation.html",
        income=income,
        expense=expense,
        budget=budget,
        savings=savings
    )
@app.route("/ai-insights")
def ai_insights():

    if "user_id" not in session:
        return redirect("/")

    conn = sqlite3.connect("database.db")
    cursor = conn.cursor()

    user_id = session["user_id"]

    # Total Income
    cursor.execute(
        "SELECT IFNULL(SUM(amount),0) FROM income WHERE user_id=?",
        (user_id,)
    )
    income = cursor.fetchone()[0]

    # Total Expense
    cursor.execute(
        "SELECT IFNULL(SUM(amount),0) FROM expenses WHERE user_id=?",
        (user_id,)
    )
    expense = cursor.fetchone()[0]

    # Total Budget
    cursor.execute(
        "SELECT IFNULL(SUM(amount),0) FROM budgets WHERE user_id=?",
        (user_id,)
    )
    budget = cursor.fetchone()[0]

    # Investment
    cursor.execute("""
        SELECT
        IFNULL(SUM(invested_amount),0),
        IFNULL(SUM(current_value),0)
        FROM investments
        WHERE user_id=?
    """,(user_id,))

    invest = cursor.fetchone()

    invested = invest[0]
    current = invest[1]

    profit = current - invested

    savings = income - expense

    insights = []

    if savings > income*0.2:
        insights.append("✅ Excellent! You are saving more than 20% of your income.")
    else:
        insights.append("💡 Try to save at least 20% of your monthly income.")

    if expense > income*0.8:
        insights.append("⚠️ Your expenses are very high.")
    else:
        insights.append("✅ Your expenses are under control.")

    if budget > 0:
        if expense > budget:
            insights.append("🚨 Budget limit exceeded.")
        else:
            insights.append("✅ You are within your budget.")

    if profit > 0:
        insights.append(f"📈 Your investments earned ₹{profit}.")
    elif profit < 0:
        insights.append(f"📉 Investment loss ₹{abs(profit)}.")
    else:
        insights.append("➖ No investment profit or loss.")

    conn.close()

    return render_template(
        "ai_insights.html",
        income=income,
        expense=expense,
        savings=savings,
        budget=budget,
        invested=invested,
        current=current,
        profit=profit,
        insights=insights
    )
@app.route("/expense-analytics")
def expense_analytics():

    if "user_id" not in session:
        return redirect("/")

    conn = sqlite3.connect("database.db")
    cursor = conn.cursor()

    user_id = session["user_id"]

    # Pie Chart
    cursor.execute("""
        SELECT expense_name,
               SUM(amount)
        FROM expenses
        WHERE user_id=?
        GROUP BY expense_name
    """,(user_id,))

    expenses = cursor.fetchall()

    # Monthly Trend
    cursor.execute("""
        SELECT strftime('%m', date),
               SUM(amount)
        FROM expenses
        WHERE user_id=?
        GROUP BY strftime('%m', date)
        ORDER BY strftime('%m', date)
    """,(user_id,))

    trend = cursor.fetchall()

    conn.close()

    return render_template(
        "expense_analytics.html",
        expenses=expenses,
        trend=trend
    )
@app.route("/reports")
def reports():

    if "user_id" not in session:
        return redirect("/login")

    conn = sqlite3.connect("database.db")
    cursor = conn.cursor()

    user_id = session["user_id"]

    # Total Income
    cursor.execute("SELECT IFNULL(SUM(amount),0) FROM income WHERE user_id=?", (user_id,))
    total_income = cursor.fetchone()[0]

    # Total Expense
    cursor.execute("SELECT IFNULL(SUM(amount),0) FROM expenses WHERE user_id=?", (user_id,))
    total_expense = cursor.fetchone()[0]

    # Download History
    cursor.execute("""
        SELECT report_type, download_date
        FROM download_history
        WHERE user_id=?
        ORDER BY id DESC
    """, (user_id,))

    history = cursor.fetchall()
    

    conn.close()

    return render_template(
        "reports.html",
        total_income=total_income,
        total_expense=total_expense,
        history=history
    )
@app.route("/download_pdf")
def download_pdf():

    user_id = session.get("user_id")

    if not user_id:
        return redirect("/")

    conn = sqlite3.connect("database.db")
    cursor = conn.cursor()

    # ---------- Summary ----------
    cursor.execute("SELECT SUM(amount) FROM income WHERE user_id=?", (user_id,))
    total_income = cursor.fetchone()[0] or 0

    cursor.execute("SELECT SUM(amount) FROM expenses WHERE user_id=?", (user_id,))
    total_expense = cursor.fetchone()[0] or 0

    cursor.execute("SELECT SUM(amount) FROM budgets WHERE user_id=?", (user_id,))
    total_budget = cursor.fetchone()[0] or 0

    savings = total_income - total_expense

    # ---------- Income ----------
    cursor.execute("""
        SELECT income_name, amount, date
        FROM income
        WHERE user_id=?
        ORDER BY date DESC
    """, (user_id,))
    income_data = cursor.fetchall()

    # ---------- Expenses ----------
    cursor.execute("""
        SELECT expense_name, amount, date, category
        FROM expenses
        WHERE user_id=?
        ORDER BY date DESC
    """, (user_id,))
    expense_data = cursor.fetchall()

    # ---------- Budget ----------
    cursor.execute("""
        SELECT budget_name, amount
        FROM budgets
        WHERE user_id=?
    """, (user_id,))
    budget_data = cursor.fetchall()

    # ---------- Goals ----------
    cursor.execute("""
        SELECT goal_name, target_amount, saved_amount
        FROM goals
        WHERE user_id=?
    """, (user_id,))
    goal_data = cursor.fetchall()
    from datetime import datetime

    cursor.execute(
        "INSERT INTO download_history(user_id, report_type, download_date) VALUES (?, ?, ?)",
        (session["user_id"], "PDF", datetime.now().strftime("%d-%m-%Y %H:%M"))
   )
    conn.commit()
    
    print("PDF history inserted")



    # ---------- Investments ----------
    cursor.execute("""
        SELECT investment_name,
               invested_amount,
               current_value
        FROM investments
        WHERE user_id=?
    """, (user_id,))
    investment_data = cursor.fetchall()

    conn.close()

    pdf_file = "Financial_Report.pdf"

    doc = SimpleDocTemplate(
        pdf_file,
        pagesize=letter
    )

    styles = getSampleStyleSheet()

    elements = []

    elements.append(
        Paragraph(
            "<b><font size='20'>FinSight Financial Report</font></b>",
            styles["Title"]
        )
    )

    elements.append(
        Paragraph(
            "Developed by Sneha",
            styles["Normal"]
        )
    )

    elements.append(Spacer(1,20))

    summary = [
        ["Description","Amount"],
        ["Total Income",f"Rs. {total_income}"],
        ["Total Expense",f"Rs. {total_expense}"],
        ["Total Budget",f"Rs. {total_budget}"],
        ["Total Savings",f"Rs. {savings}"]
    ]

    table = Table(summary)

    table.setStyle(TableStyle([
        ("BACKGROUND",(0,0),(-1,0),colors.blue),
        ("TEXTCOLOR",(0,0),(-1,0),colors.white),
        ("GRID",(0,0),(-1,-1),1,colors.black),
        ("BACKGROUND",(0,1),(-1,-1),colors.beige),
        ("FONTNAME",(0,0),(-1,0),"Helvetica-Bold"),
        ("ALIGN",(0,0),(-1,-1),"CENTER"),
    ]))

    elements.append(table)
        # ================= Income Details =================

    elements.append(Spacer(1,20))
    elements.append(Paragraph("<b>Income Details</b>", styles["Heading2"]))

    income_table = [["Income Name", "Amount", "Date"]]

    for i in income_data:
        income_table.append([
            i[0],
            f"Rs. {i[1]}",
            i[2]
        ])

    table = Table(income_table)

    table.setStyle(TableStyle([
        ("BACKGROUND",(0,0),(-1,0),colors.darkblue),
        ("TEXTCOLOR",(0,0),(-1,0),colors.white),
        ("GRID",(0,0),(-1,-1),1,colors.black),
        ("BACKGROUND",(0,1),(-1,-1),colors.whitesmoke),
        ("FONTNAME",(0,0),(-1,0),"Helvetica-Bold"),
        ("ALIGN",(0,0),(-1,-1),"CENTER"),
    ]))

    elements.append(table)


    # ================= Expense Details =================

    elements.append(Spacer(1,20))
    elements.append(Paragraph("<b>Expense Details</b>", styles["Heading2"]))

    expense_table = [["Expense Name", "Amount", "Date", "Category"]]

    for e in expense_data:
        expense_table.append([
            e[0],
            f"Rs. {e[1]}",
            e[2],
            e[3]
        ])

    table = Table(expense_table)

    table.setStyle(TableStyle([
        ("BACKGROUND",(0,0),(-1,0),colors.green),
        ("TEXTCOLOR",(0,0),(-1,0),colors.white),
        ("GRID",(0,0),(-1,-1),1,colors.black),
        ("BACKGROUND",(0,1),(-1,-1),colors.whitesmoke),
        ("FONTNAME",(0,0),(-1,0),"Helvetica-Bold"),
        ("ALIGN",(0,0),(-1,-1),"CENTER"),
    ]))

    elements.append(table)
        # ================= Budget Details =================

    elements.append(Spacer(1,20))
    elements.append(Paragraph("<b>Budget Details</b>", styles["Heading2"]))

    budget_table = [["Budget Name", "Amount"]]

    for b in budget_data:
        budget_table.append([
            b[0],
            f"Rs. {b[1]}"
        ])

    table = Table(budget_table)

    table.setStyle(TableStyle([
        ("BACKGROUND", (0,0), (-1,0), colors.orange),
        ("TEXTCOLOR", (0,0), (-1,0), colors.white),
        ("GRID", (0,0), (-1,-1), 1, colors.black),
        ("BACKGROUND", (0,1), (-1,-1), colors.whitesmoke),
        ("FONTNAME", (0,0), (-1,0), "Helvetica-Bold"),
        ("ALIGN", (0,0), (-1,-1), "CENTER"),
    ]))

    elements.append(table)


    # ================= Goal Details =================

    elements.append(Spacer(1,20))
    elements.append(Paragraph("<b>Goal Details</b>", styles["Heading2"]))

    goal_table = [["Goal Name", "Target Amount", "Saved Amount"]]

    for g in goal_data:
        goal_table.append([
            g[0],
            f"Rs. {g[1]}",
            f"Rs. {g[2]}"
        ])

    table = Table(goal_table)

    table.setStyle(TableStyle([
        ("BACKGROUND", (0,0), (-1,0), colors.purple),
        ("TEXTCOLOR", (0,0), (-1,0), colors.white),
        ("GRID", (0,0), (-1,-1), 1, colors.black),
        ("BACKGROUND", (0,1), (-1,-1), colors.whitesmoke),
        ("FONTNAME", (0,0), (-1,0), "Helvetica-Bold"),
        ("ALIGN", (0,0), (-1,-1), "CENTER"),
    ]))

    elements.append(table)
        # ================= Investment Details =================

    elements.append(Spacer(1,20))
    elements.append(Paragraph("<b>Investment Details</b>", styles["Heading2"]))

    investment_table = [["Investment Name", "Invested Amount", "Current Value"]]

    for inv in investment_data:
        investment_table.append([
            inv[0],
            f"Rs. {inv[1]}",
            f"Rs. {inv[2]}"
        ])

    table = Table(investment_table)

    table.setStyle(TableStyle([
        ("BACKGROUND", (0,0), (-1,0), colors.darkgreen),
        ("TEXTCOLOR", (0,0), (-1,0), colors.white),
        ("GRID", (0,0), (-1,-1), 1, colors.black),
        ("BACKGROUND", (0,1), (-1,-1), colors.whitesmoke),
        ("FONTNAME", (0,0), (-1,0), "Helvetica-Bold"),
        ("ALIGN", (0,0), (-1,-1), "CENTER"),
    ]))

    elements.append(table)


    # ================= Generate PDF =================

    doc.build(elements)

    return send_file(
        pdf_file,
        as_attachment=True,
        download_name="Financial_Report.pdf"
    )
@app.route("/download_excel")
def download_excel():

    user_id = session.get("user_id")

    if not user_id:
        return redirect("/")

    conn = sqlite3.connect("database.db")
    cursor = conn.cursor()

    wb = Workbook()

    # ================= SUMMARY =================
    ws = wb.active
    ws.title = "Summary"

    cursor.execute("SELECT SUM(amount) FROM income WHERE user_id=?", (user_id,))
    total_income = cursor.fetchone()[0] or 0

    cursor.execute("SELECT SUM(amount) FROM expenses WHERE user_id=?", (user_id,))
    total_expense = cursor.fetchone()[0] or 0

    cursor.execute("SELECT SUM(amount) FROM budgets WHERE user_id=?", (user_id,))
    total_budget = cursor.fetchone()[0] or 0
    from datetime import datetime

    cursor.execute(
        "INSERT INTO download_history(user_id, report_type, download_date) VALUES (?, ?, ?)",
        (session["user_id"], "Excel", datetime.now().strftime("%d-%m-%Y %H:%M"))
   )
    conn.commit()
    
    print("Excel history inserted")

    savings = total_income - total_expense

    ws.append(["Description", "Amount"])
    ws.append(["Total Income", total_income])
    ws.append(["Total Expense", total_expense])
    ws.append(["Total Budget", total_budget])
    ws.append(["Total Savings", savings])

    # ================= INCOME =================
    ws2 = wb.create_sheet("Income")

    ws2.append(["Income Name", "Amount", "Date"])

    cursor.execute("""
        SELECT income_name, amount, date
        FROM income
        WHERE user_id=?
    """, (user_id,))

    for row in cursor.fetchall():
        ws2.append(row)

    # ================= EXPENSE =================
    ws3 = wb.create_sheet("Expenses")

    ws3.append(["Expense Name", "Amount", "Date", "Category"])

    cursor.execute("""
        SELECT expense_name, amount, date, category
        FROM expenses
        WHERE user_id=?
    """, (user_id,))

    for row in cursor.fetchall():
        ws3.append(row)

    # ================= BUDGET =================
    ws4 = wb.create_sheet("Budgets")

    ws4.append(["Budget Name", "Amount"])

    cursor.execute("""
        SELECT budget_name, amount
        FROM budgets
        WHERE user_id=?
    """, (user_id,))

    for row in cursor.fetchall():
        ws4.append(row)

    # ================= GOALS =================
    ws5 = wb.create_sheet("Goals")

    ws5.append(["Goal Name", "Target Amount", "Saved Amount"])

    cursor.execute("""
        SELECT goal_name, target_amount, saved_amount
        FROM goals
        WHERE user_id=?
    """, (user_id,))

    for row in cursor.fetchall():
        ws5.append(row)

    # ================= INVESTMENTS =================
    ws6 = wb.create_sheet("Investments")

    ws6.append(["Investment", "Invested Amount", "Current Value"])

    cursor.execute("""
        SELECT investment_name, invested_amount, current_value
        FROM investments
        WHERE user_id=?
    """, (user_id,))

    for row in cursor.fetchall():
        ws6.append(row)

    file_name = "Financial_Report.xlsx"

    wb.save(file_name)

    conn.close()

    return send_file(file_name, as_attachment=True)
@app.route("/profile")
def profile():

    user_id = session.get("user_id")

    if not user_id:
        return redirect("/")

    conn = sqlite3.connect("database.db")
    cursor = conn.cursor()

    cursor.execute("""
        SELECT *
        FROM users
        WHERE id=?
    """, (user_id,))

    user = cursor.fetchone()

    conn.close()

    return render_template(
        "profile.html",
        user=user
    )


@app.route("/update_profile", methods=["POST"])
def update_profile():

    user_id = session.get("user_id")

    if not user_id:
        return redirect("/")

    name = request.form["name"]
    email = request.form["email"]

    conn = sqlite3.connect("database.db")
    cursor = conn.cursor()

    cursor.execute("""
        UPDATE users
        SET name=?,
            email=?
        WHERE id=?
    """, (
        name,
        email,
        user_id
    ))

    conn.commit()
    conn.close()

    flash("Profile Updated Successfully!")

    return redirect("/profile")
@app.route("/change_password", methods=["POST"])
def change_password():

    user_id = session.get("user_id")

    if not user_id:
        return redirect("/")

    current_password = request.form["current_password"]
    new_password = request.form["new_password"]

    conn = sqlite3.connect("database.db")
    cursor = conn.cursor()

    cursor.execute(
        "SELECT password FROM users WHERE id=?",
        (user_id,)
    )

    user = cursor.fetchone()

    if not user:
        conn.close()
        flash("User not found!")
        return redirect("/profile")

    if user[0] != current_password:
        conn.close()
        flash("Current Password is incorrect!")
        return redirect("/profile")

    cursor.execute(
        "UPDATE users SET password=? WHERE id=?",
        (new_password, user_id)
    )

    conn.commit()
    conn.close()

    flash("Password Changed Successfully!")

    return redirect("/profile")


@app.route("/logout")
def logout():

    session.clear()

    flash("Logged Out Successfully!")

    return redirect("/")


if __name__ == "__main__":
    app.run(debug=True)